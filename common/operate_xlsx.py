import openpyxl
import pandas as pd
from common.get_path import api_xlsx
from common import excel_font_conf

xlsx_path = api_xlsx()


class Excel:
    def __init__(self):
        self.dataframe = pd.read_excel(xlsx_path)
        self.excel = openpyxl.load_workbook(xlsx_path)
        self.sheet = self.excel['Sheet1']

    def get_row(self, row):
        return self.dataframe.iloc[row - 1]

    def get_col(self, col):
        return self.dataframe[col]

    def get_cell(self, row, col):
        return self.dataframe.iloc[row, col]

    def get_value_row_id(self, value1, value2):
        excel_index_list = []
        df_index_list = self.dataframe.loc[self.dataframe[value1] == value2].index
        for index in df_index_list:
            excel_index_list.append(index + 1)

        return excel_index_list

    def write_pass(self, row):
        excel_font_conf.pass_font(self.sheet.cell(row + 1, 9))
        self.excel.save(xlsx_path)

    def write_fail(self, row):
        excel_font_conf.fail_font(self.sheet.cell(row + 1, 9))
        self.excel.save(xlsx_path)


if __name__ == '__main__':
    a = Excel()
    # print(a.dataframe)
    # print(a.get_row(1))
    # print(a.get_col("所属模块"))
    # print(a.get_cell(1, 3))
    # print(a.get_value_row_id("所属模块", "账号相关|登录"))
    # a.write_pass(2, 9)
    # a.write_fail(3, 9)
