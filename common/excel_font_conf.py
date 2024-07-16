from openpyxl.styles import PatternFill, Font


# pass的写入配置
def pass_font(cell):
    # 定义写入的内容
    cell.value = 'PASS'
    # 定义单元格的格式颜色和样式，绿色加粗
    cell.fill = PatternFill('solid', fgColor='AACF91')
    cell.font = Font(bold=True)


# failed的写入配置
def fail_font(cell):
    # 定义写入的内容
    cell.value = 'FAILED'
    # 定义单元格的格式颜色和样式，红色加粗
    cell.fill = PatternFill('solid', fgColor='FF0000')
    cell.font = Font(bold=True)
